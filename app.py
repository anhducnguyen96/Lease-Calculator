import io
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st

st.set_page_config(page_title="Lease Overview App", layout="wide")
st.title("🏢 Multi-Tab Lease Overview App")

st.markdown("""
Upload an Excel workbook containing the necessary sheets:
* **`Square_Meters`** / `Square_Metres`
* **`Prices`**
* **`Add-on`**
* **`Add-on participation`**
""")

uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx", "xls"])


def format_cz_number(val):
    """Formats numeric floats into Czech number format string for Streamlit UI display only."""
    if pd.isnull(val):
        return ""
    try:
        val_float = float(val)
        if abs(val_float) < 0.001:
            return ""
        return (
            f"{val_float:,.2f}"
            .replace(",", "X")
            .replace(".", ",")
            .replace("X", " ")
        )
    except (ValueError, TypeError):
        return str(val)


def add_total_row_ui(df, label_col, numeric_cols, total_label="Celkem"):
    """Appends a static 'Celkem' sum row at the bottom for Streamlit web display."""
    df_with_total = df.copy()
    total_data = {col: "" for col in df.columns}
    total_data[label_col] = total_label

    for col in numeric_cols:
        if col in df.columns:
            total_data[col] = df[col].sum()

    df_total_row = pd.DataFrame([total_data])
    return pd.concat([df_with_total, df_total_row], ignore_index=True)


def get_column_configs(df, numeric_cols):
    """Generates column configuration to right-align numeric columns in Streamlit."""
    configs = {}
    for col in df.columns:
        if col in numeric_cols:
            configs[col] = st.column_config.TextColumn(col, alignment="right")
        else:
            configs[col] = st.column_config.TextColumn(col, alignment="left")
    return configs


def process_original_lease_data(df_raw):
    """Processes raw lease DataFrame, coercing numbers, suppressing empty rows, and formatting for UI."""
    df = df_raw.copy().dropna(how="all", axis=1).dropna(how="all", axis=0)

    numeric_cols = list(df.columns[1:])
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    # Filter out rows where all numeric values are 0
    df = df[~(df[numeric_cols] == 0).all(axis=1)].reset_index(drop=True)

    # UI Total Row & Display formatting
    label_col = df.columns[0]
    df_with_total = add_total_row_ui(df, label_col, numeric_cols, total_label="Celkem")

    df_display = df_with_total.copy()
    for col in numeric_cols:
        df_display[col] = df_display[col].apply(format_cz_number)

    configs = get_column_configs(df_display, numeric_cols)
    return df, df_display, configs, numeric_cols


def process_price_list_data(df_raw):
    """Formats numeric columns in tabular data for UI display."""
    df_prices_display = df_raw.copy()
    numeric_cols = []

    for col in df_prices_display.columns:
        if pd.api.types.is_numeric_dtype(df_prices_display[col]) or any(
            isinstance(val, (int, float)) for val in df_prices_display[col].dropna()
        ):
            numeric_cols.append(col)
            df_prices_display[col] = df_prices_display[col].apply(format_cz_number)

    configs = get_column_configs(df_prices_display, numeric_cols)
    return df_prices_display, configs


def generate_excel_download_with_formulas(raw_panels):
    """
    Builds a native openpyxl Excel workbook:
    - Bold & Center Aligned headers and 'Celkem' rows.
    - Gridlines removed on all sheets.
    - Automatic column widths and row heights.
    - Native Excel number formatting '#,##0.00'.
    - Dynamic Excel =SUM(col_start:col_end) formulas.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    cz_num_format = "#,##0.00"
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold_center_align = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)

    for sheet_name, data in raw_panels.items():
        ws = wb.create_sheet(title=sheet_name)

        # Hide Gridlines
        ws.views.sheetView[0].showGridLines = False

        df = data["df"].copy()
        add_total = data.get("add_total", False)

        # Write Header
        headers = list(df.columns)
        ws.append(headers)

        # Style Header Row
        ws.row_dimensions[1].height = 28
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_alignment

        # Write Raw Data Rows
        for _, row in df.iterrows():
            row_vals = []
            for col in headers:
                val = row[col]
                if isinstance(val, (int, float)):
                    row_vals.append(val)
                elif pd.isna(val):
                    row_vals.append("")
                else:
                    row_vals.append(val)
            ws.append(row_vals)

        last_data_row = ws.max_row

        # Add native Excel =SUM(...) row if required
        if add_total and last_data_row >= 2:
            total_row = ["Celkem"]
            for col_idx in range(2, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                total_row.append(f"=SUM({col_letter}2:{col_letter}{last_data_row})")
            ws.append(total_row)

            # Style Total Row (Bold & Center Aligned)
            total_row_idx = ws.max_row
            ws.row_dimensions[total_row_idx].height = 22
            for cell in ws[total_row_idx]:
                cell.font = bold_font
                cell.alignment = bold_center_align

        # Format Numbers & Row Heights
        for r_idx in range(2, ws.max_row + 1):
            if not (add_total and r_idx == ws.max_row):
                ws.row_dimensions[r_idx].height = 20

            for c_idx in range(2, len(headers) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                    cell.number_format = cz_num_format

        # Automatic Column Widths Adjustment
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or "")
                if cell.number_format == cz_num_format and isinstance(cell.value, (int, float)):
                    val_str = f"{cell.value:,.2f}"
                max_len = max(max_len, len(val_str))

            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


if uploaded_file is not None:
    try:
        xls = pd.ExcelFile(uploaded_file)
        sheet_names = xls.sheet_names

        sqm_sheet = next((s for s in sheet_names if "square" in s.lower()), sheet_names[0])
        prices_sheet = next((s for s in sheet_names if "price" in s.lower()), None)
        addon_sheet = next((s for s in sheet_names if "add-on" in s.lower() and "participation" not in s.lower()), None)
        part_sheet = next((s for s in sheet_names if "participation" in s.lower()), None)

        df_sqm_raw = pd.read_excel(uploaded_file, sheet_name=sqm_sheet)

        raw_export_panels = {}

        tab1, tab2, tab3, tab4, tab5 = st.tabs(
            [
                "📋 Panel 1: Main Lease List",
                "🏷️ Panel 2: Price List",
                "➕ Panel 3: Add-on Rates",
                "🤝 Panel 4: Participation Rules",
                "🧮 Panel 5: Calculated Lease Summary",
            ]
        )

        # ------------------------------------------------------------------
        # PANEL 1: Main Lease List (Raw Quantities)
        # ------------------------------------------------------------------
        with tab1:
            st.subheader(f"Main Lease Overview (Sheet: `{sqm_sheet}`)")
            df_sqm_clean, df_sqm_display, configs1, num_cols1 = process_original_lease_data(df_sqm_raw)
            raw_export_panels["Panel 1 - Main Lease"] = {"df": df_sqm_clean, "add_total": True}
            st.dataframe(
                df_sqm_display,
                width="stretch",
                height="content",
                hide_index=True,
                column_config=configs1
            )

        # ------------------------------------------------------------------
        # PANEL 2: Master Price List
        # ------------------------------------------------------------------
        with tab2:
            st.subheader(f"Master Prices per Unit " + (f"(Sheet: `{prices_sheet}`)" if prices_sheet else ""))
            if prices_sheet:
                df_prices_raw = pd.read_excel(uploaded_file, sheet_name=prices_sheet)
                df_prices_raw.columns = df_prices_raw.columns.str.strip()
                df_prices_display, configs2 = process_price_list_data(df_prices_raw)
                raw_export_panels["Panel 2 - Prices"] = {"df": df_prices_raw, "add_total": False}
                st.dataframe(
                    df_prices_display,
                    width="stretch",
                    height="content",
                    hide_index=True,
                    column_config=configs2
                )
            else:
                st.warning("No sheet found for Prices.")

        # ------------------------------------------------------------------
        # PANEL 3: Add-on Rates
        # ------------------------------------------------------------------
        addon_rates = {}
        with tab3:
            st.subheader(f"Add-on Percentage Rates " + (f"(Sheet: `{addon_sheet}`)" if addon_sheet else ""))
            if addon_sheet:
                df_addon_raw = pd.read_excel(uploaded_file, sheet_name=addon_sheet)
                df_addon_raw.columns = df_addon_raw.columns.str.strip()

                if "Typ" in df_addon_raw.columns and "Add-on" in df_addon_raw.columns:
                    for _, row in df_addon_raw.iterrows():
                        key = str(row["Typ"]).strip()
                        rate = pd.to_numeric(row["Add-on"], errors="coerce")
                        if pd.notnull(rate):
                            addon_rates[key] = float(rate)

                df_addon_display, configs_addon = process_price_list_data(df_addon_raw)
                raw_export_panels["Panel 3 - Add-on Rates"] = {"df": df_addon_raw, "add_total": False}
                st.dataframe(
                    df_addon_display,
                    width="stretch",
                    height="content",
                    hide_index=True,
                    column_config=configs_addon
                )
            else:
                st.warning("No sheet found for Add-on rates.")

        # ------------------------------------------------------------------
        # PANEL 4: Add-on Participation Table
        # ------------------------------------------------------------------
        df_part_raw = None
        with tab4:
            st.subheader(f"Add-on Participation Matrix " + (f"(Sheet: `{part_sheet}`)" if part_sheet else ""))
            if part_sheet:
                df_part_raw = pd.read_excel(uploaded_file, sheet_name=part_sheet)
                df_part_raw.columns = df_part_raw.columns.str.strip()
                raw_export_panels["Panel 4 - Participation"] = {"df": df_part_raw, "add_total": False}
                st.dataframe(
                    df_part_raw,
                    width="stretch",
                    height="content",
                    hide_index=True
                )
            else:
                st.warning("No sheet found for Add-on participation.")

        # ------------------------------------------------------------------
        # PANEL 5: Calculated Lease Summary
        # ------------------------------------------------------------------
        with tab5:
            st.subheader("Calculated Lease Totals (CZK) - Conditional Participation Included")

            if prices_sheet and "Kód" in df_prices_raw.columns and "Cena za jednotku (Bez DPH)" in df_prices_raw.columns:
                df_prices_base = df_prices_raw.drop_duplicates(subset=["Kód"], keep="first")
                price_map = dict(zip(df_prices_base["Kód"], df_prices_base["Cena za jednotku (Bez DPH)"]))

                def check_participation(company, typ_str, category):
                    if df_part_raw is None or df_part_raw.empty:
                        return True
                    match = df_part_raw[
                        (df_part_raw["Společnost"].astype(str).str.strip() == str(company).strip()) &
                        (df_part_raw["Typ"].astype(str).str.strip().str.lower() == str(typ_str).strip().lower())
                    ]
                    if not match.empty and category in match.columns:
                        val = str(match[category].values[0]).strip().upper()
                        return val in ["YES", "ANO", "TRUE", "1"]
                    return False

                label_col = df_sqm_clean.columns[0]
                calc_dict = {label_col: df_sqm_clean[label_col]}
                calc_cols = []

                for col in num_cols1:
                    if col in price_map:
                        unit_price = price_map[col]

                        calculated_series = []
                        for idx, row in df_sqm_clean.iterrows():
                            company = row[label_col]
                            sqm_val = row[col]

                            addon_pct = 0.0

                            if col.startswith("K - Typ"):
                                typ_name = col.replace("K - ", "").strip()
                                addon_pct += addon_rates.get("Výtah, schodiště, lobby", 0.0)

                                if check_participation(company, typ_name, "JM interní"):
                                    addon_pct += addon_rates.get("JM interní", 0.0)
                                if check_participation(company, typ_name, "JM reprezentativní"):
                                    addon_pct += addon_rates.get("JM reprezentativní", 0.0)
                                if check_participation(company, typ_name, "Chodby, kuchyň, copy"):
                                    addon_pct += addon_rates.get("Chodby, kuchyň, copy", 0.0)

                            elif col == "Terasy":
                                if check_participation(company, "Terasy", "Terasy") or check_participation(company, "Typ A", "Terasy"):
                                    addon_pct += addon_rates.get("Terasy", 0.0)

                            elif col == "Sklad":
                                addon_pct += addon_rates.get("Sklad", 0.0)

                            item_total = sqm_val * (1.0 + addon_pct) * unit_price
                            calculated_series.append(item_total)

                        calc_dict[col] = calculated_series
                        calc_cols.append(col)

                df_calc = pd.DataFrame(calc_dict)

                if calc_cols:
                    df_calc["Celková Cena (CZK)"] = df_calc[calc_cols].sum(axis=1)

                if "Celková Cena (CZK)" in df_calc.columns:
                    df_calc = df_calc[df_calc["Celková Cena (CZK)"] > 0].reset_index(drop=True)

                raw_export_panels["Panel 5 - Calculated Summary"] = {"df": df_calc, "add_total": True}

                # UI Display with Total Row
                numeric_calc_cols = [c for c in df_calc.columns if c != label_col]
                df_calc_with_total = add_total_row_ui(df_calc, label_col, numeric_calc_cols, total_label="Celkem")
                df_calc_display, configs5 = process_price_list_data(df_calc_with_total)

                st.dataframe(
                    df_calc_display,
                    width="stretch",
                    height="content",
                    hide_index=True,
                    column_config=configs5
                )
            else:
                st.info("Panel 5 will display calculated totals once all required sheets and columns are available.")

        # ------------------------------------------------------------------
        # EXCEL DOWNLOAD BUTTON
        # ------------------------------------------------------------------
        if raw_export_panels:
            st.divider()
            excel_data = generate_excel_download_with_formulas(raw_export_panels)
            st.download_button(
                label="📥 Download Full Report as Excel Workbook",
                data=excel_data,
                file_name="Lease_Overview_Export.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"Error reading workbook: {e}")
